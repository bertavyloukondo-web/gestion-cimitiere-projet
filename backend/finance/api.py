from ninja import Router, Schema
from typing import Optional, List
from finance.models import Facture, Paiement
from django.http import HttpResponse
import datetime
import random

router = Router()


class FactureSchema(Schema):
    id: int
    reservation_id: int
    client_id: int
    numero: str
    montant_total: float
    montant_paye: float
    statut: str
    date_emission: datetime.datetime
    date_echeance: datetime.date


class FactureCreateSchema(Schema):
    reservation_id: int
    client_id: int
    montant_total: float
    date_echeance: datetime.date


class PaiementSchema(Schema):
    id: int
    facture_id: int
    montant: float
    mode_paiement: str
    reference: Optional[str] = None
    date_paiement: datetime.datetime


class PaiementCreateSchema(Schema):
    facture_id: int
    montant: float
    mode_paiement: str
    reference: Optional[str] = None


@router.get("/factures", response=List[FactureSchema])
def list_factures(request, statut: Optional[str] = None):
    qs = Facture.objects.all()
    if statut:
        qs = qs.filter(statut=statut)
    return qs


@router.post("/factures", response=FactureSchema)
def create_facture(request, data: FactureCreateSchema):
    numero = f"FAC-{datetime.date.today().year}-{random.randint(1000, 9999)}"
    facture = Facture.objects.create(numero=numero, **data.dict())
    return facture


@router.post("/paiements", response=PaiementSchema)
def create_paiement(request, data: PaiementCreateSchema):
    facture = Facture.objects.get(id=data.facture_id)
    paiement = Paiement.objects.create(**data.dict())
    facture.montant_paye = float(facture.montant_paye) + float(data.montant)
    if float(facture.montant_paye) >= float(facture.montant_total):
        facture.statut = 'payee'
    else:
        facture.statut = 'partiellement_payee'
    facture.save()
    return paiement


@router.get("/paiements", response=List[PaiementSchema])
def list_paiements(request, mode_paiement: Optional[str] = None):
    qs = Paiement.objects.all().order_by('-date_paiement')
    if mode_paiement:
        qs = qs.filter(mode_paiement=mode_paiement)
    return qs


@router.get("/stats")
def stats_financieres(request):
    total_factures = Facture.objects.count()
    total_paye = sum(f.montant_paye for f in Facture.objects.all())
    total_attendu = sum(f.montant_total for f in Facture.objects.all())
    return {
        "total_factures": total_factures,
        "total_paye": float(total_paye),
        "total_attendu": float(total_attendu),
    }


@router.get("/factures/{facture_id}/pdf")
def generer_pdf(request, facture_id: int):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import cm
    import io

    try:
        facture = Facture.objects.get(id=facture_id)
        client = facture.client
        reservation = facture.reservation

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        styles = getSampleStyleSheet()
        elements = []

        elements.append(Paragraph("FACTURE", styles['Title']))
        elements.append(Spacer(1, 0.5*cm))

        elements.append(Paragraph(f"<b>Numero :</b> {facture.numero}", styles['Normal']))
        elements.append(Paragraph(f"<b>Date emission :</b> {facture.date_emission.strftime('%d/%m/%Y')}", styles['Normal']))
        elements.append(Paragraph(f"<b>Date echeance :</b> {facture.date_echeance}", styles['Normal']))
        elements.append(Paragraph(f"<b>Statut :</b> {facture.statut.upper()}", styles['Normal']))
        elements.append(Spacer(1, 0.5*cm))

        elements.append(Paragraph("<b>INFORMATIONS CLIENT</b>", styles['Heading2']))
        elements.append(Paragraph(f"Nom : {client.username}", styles['Normal']))
        elements.append(Paragraph(f"Email : {client.email}", styles['Normal']))
        elements.append(Spacer(1, 0.5*cm))

        elements.append(Paragraph("<b>INFORMATIONS DEFUNT</b>", styles['Heading2']))
        elements.append(Paragraph(f"Nom : {reservation.nom_defunt} {reservation.prenom_defunt}", styles['Normal']))
        elements.append(Paragraph(f"Date de naissance : {reservation.date_naissance_defunt}", styles['Normal']))
        elements.append(Paragraph(f"Date de deces : {reservation.date_deces_defunt}", styles['Normal']))
        elements.append(Spacer(1, 0.5*cm))

        elements.append(Paragraph("<b>DETAILS FINANCIERS</b>", styles['Heading2']))
        data = [
            ["Description", "Montant"],
            ["Concession funeraire", f"{facture.montant_total} FCFA"],
            ["Montant paye", f"{facture.montant_paye} FCFA"],
            ["Reste a payer", f"{float(facture.montant_total) - float(facture.montant_paye)} FCFA"],
        ]
        table = Table(data, colWidths=[12*cm, 5*cm])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1A237E')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#F5F5F5')),
        ]))
        elements.append(table)
        elements.append(Spacer(1, 1*cm))
        elements.append(Paragraph("Merci pour votre confiance.", styles['Normal']))

        doc.build(elements)
        buffer.seek(0)
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="facture_{facture.numero}.pdf"'
        return response

    except Facture.DoesNotExist:
        return {"error": "Facture introuvable"}


@router.get("/export-excel")
def export_excel(request):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    import io

    paiements = Paiement.objects.all().select_related('facture').order_by('-date_paiement')

    wb = Workbook()
    sheet = wb.active
    sheet.title = "Transactions"

    headers = ["N° Facture", "Montant (FCFA)", "Mode de paiement", "Référence", "Date de paiement"]
    sheet.append(headers)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", start_color="1A237E")
    for col in range(1, len(headers) + 1):
        cell = sheet.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    mode_labels = {
        "mobile_money": "Mobile Money",
        "airtel_money": "Airtel Money",
        "especes": "Espèces",
        "virement": "Virement",
    }

    for p in paiements:
        sheet.append([
            p.facture.numero,
            float(p.montant),
            mode_labels.get(p.mode_paiement, p.mode_paiement),
            p.reference or "-",
            p.date_paiement.strftime("%d/%m/%Y %H:%M"),
        ])

    for col, width in zip("ABCDE", [18, 18, 20, 25, 20]):
        sheet.column_dimensions[col].width = width

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="transactions_comptabilite.xlsx"'
    return response